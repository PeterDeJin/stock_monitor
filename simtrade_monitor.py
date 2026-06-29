import shioaji as sj
import time
import requests
import urllib.parse
from datetime import datetime
import os
import pytz
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ================= 配置設定 =================
API_KEY    = os.environ.get("SHIOAJI_API_KEY",    "")
SECRET_KEY = os.environ.get("SHIOAJI_SECRET_KEY", "")
IS_SIMULATION    = os.environ.get("SHIOAJI_SIMULATION", "false").lower() == "true"
VOLUME_THRESHOLD = 100   # 試撮量門檻 (張)
LIMIT_ALERT_PCT  = 0.02  # 距漲跌停 2% 以內觸發警示
SURGE_ALERT_PCT  = 8.0   # 漲跌幅超過此值（%）時特別標注

TZ_TW = pytz.timezone("Asia/Taipei")
MARKET_CLOSE_HOUR   = 13
MARKET_CLOSE_MINUTE = 35  # 13:35 後自動結束

_bark_env = os.environ.get("BARK_KEYS", "")
BARK_KEYS = [k.strip() for k in _bark_env.split(",") if k.strip()]
MONITOR_LABEL = os.environ.get("MONITOR_LABEL", "")   # 推播標題前綴（實驗版用來和正式版區分）

last_push_time    = {}
stock_state       = {}
today_sim_records = []

api = sj.Shioaji(simulation=IS_SIMULATION)


# ─────────────── 工具函式 ─────────────────────────────────

def send_bark_alert(title: str, content: str):
    if MONITOR_LABEL:
        title = f"{MONITOR_LABEL}{title}"
    enc_title   = urllib.parse.quote(title)
    enc_content = urllib.parse.quote(content)
    for key in BARK_KEYS:
        url = f"https://api.day.app/{key}/{enc_title}/{enc_content}"
        try:
            requests.get(url, timeout=3)
        except Exception as e:
            print(f"❌ Bark 推送失敗: {e}")


def tick_type_str(tick_type: int) -> str:
    return {1: "外盤", 2: "內盤"}.get(tick_type, "不明")


def check_near_limit(price: float, limit_up, limit_down) -> str:
    if limit_up   and price >= limit_up   * (1 - LIMIT_ALERT_PCT):
        return "漲停注意"
    if limit_down and price <= limit_down * (1 + LIMIT_ALERT_PCT):
        return "跌停注意"
    return ""


def calc_change_pct(price, reference) -> float | None:
    """計算相對昨收的漲跌幅（%）。price/reference 可能是 Decimal，統一轉 float 再算。"""
    if reference is None:
        return None
    price, reference = float(price), float(reference)
    if reference > 0:
        return round((price - reference) / reference * 100, 2)
    return None


def format_change_pct(pct) -> str:
    """格式化漲跌幅，超過 SURGE_ALERT_PCT 加上警示符號"""
    if pct is None:
        return "N/A"
    sign = "+" if pct >= 0 else ""
    tag  = " 🚨大幅異動" if abs(pct) >= SURGE_ALERT_PCT else ""
    return f"{sign}{pct:.2f}%{tag}"


def format_pl(price, base) -> str:
    """以 base 為基準，回傳『金額 (百分比)』，例如 +0.85 (+3.20%)。取不到值回 N/A。"""
    if price is None or base is None or base == 0:
        return "N/A"
    diff = price - base
    pct  = diff / base * 100
    sign = "+" if diff >= 0 else "-"
    return f"{sign}{abs(diff):.2f} ({sign}{abs(pct):.2f}%)"


def _to_float(v):
    """安全轉 float；shioaji 某些欄位可能是字串或 None，避免靜默變成 None。"""
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _init_state(code: str, limit_up=None, limit_down=None, reference=None):
    stock_state[code] = {
        "last_normal_price"     : None,
        "last_normal_tick_type" : 0,
        "last_normal_total_vol" : 0,
        "in_sim"                : False,
        "sim_start_time"        : None,
        "sim_price"             : None,
        "sim_total_vol"         : 0,
        "limit_up"              : limit_up,
        "limit_down"            : limit_down,
        "reference"             : reference,   # 昨收參考價
        "near_limit"            : "",
        "change_pct"            : None,        # 試撮時的漲跌幅
        "pre_sim_price"         : None,        # 進試撮前末價
        "sim_first_price"       : None,        # 試撮後首價
        "sim_high"              : None,        # 試撮期間最高價
        "sim_low"               : None,        # 試撮期間最低價
    }


# ─────────────── Tick 處理 ────────────────────────────────

def on_tick_handler(exchange, tick):
    code     = tick.code
    time_int = tick.datetime.hour * 100 + tick.datetime.minute
    is_trading_time = (900 <= time_int < 1325)
    close    = float(tick.close)   # shioaji 回傳 Decimal，統一轉 float，避免與 float 的 reference/limit 混算

    if code not in stock_state:
        _init_state(code)

    state = stock_state[code]

    # ── 正常成交（非試撮）────────────────────────────────
    if not tick.simtrade:
        if state["in_sim"]:
            state["in_sim"] = False
            record = {
                "date"           : datetime.now(TZ_TW).strftime("%Y-%m-%d"),
                "code"           : code,
                "start_time"     : state["sim_start_time"],
                "end_time"       : tick.datetime.strftime("%H:%M:%S"),
                "pre_sim_price"  : state["pre_sim_price"],     # 進試撮前末價
                "sim_first_price": state["sim_first_price"],   # 試撮後首價
                "sim_last_price" : state["sim_price"],         # 試撮末價（最後一筆試撮）
                "end_price"      : close,                 # 結束後首價
                "change_pct"     : state["change_pct"],
                "tick_type"      : tick_type_str(state["last_normal_tick_type"]),
                "pre_total_vol"  : state["last_normal_total_vol"],
                "sim_vol"        : state["sim_total_vol"],
                "sim_high"       : state["sim_high"],
                "sim_low"        : state["sim_low"],
                "near_limit"     : state["near_limit"],
            }
            today_sim_records.append(record)
            print(f"📝 [{code}] 試撮結束 | 結束價:{close:.2f} | {record['end_time']}")

        state["last_normal_price"]     = close
        state["last_normal_tick_type"] = getattr(tick, "tick_type", 0)
        state["last_normal_total_vol"] = getattr(tick, "total_volume", 0)
        return

    # ── 試撮 (simtrade=True) ──────────────────────────────
    if not is_trading_time or tick.volume < VOLUME_THRESHOLD:
        return

    near_limit  = check_near_limit(close, state["limit_up"], state["limit_down"])
    change_pct  = calc_change_pct(close, state["reference"])
    is_surge    = change_pct is not None and abs(change_pct) >= SURGE_ALERT_PCT

    if not state["in_sim"]:
        state["in_sim"]          = True
        state["sim_start_time"]  = tick.datetime.strftime("%H:%M:%S")
        state["sim_price"]       = close
        state["sim_total_vol"]   = tick.volume
        state["near_limit"]      = near_limit
        state["change_pct"]      = change_pct
        state["pre_sim_price"]   = state["last_normal_price"]  # 進試撮前末價
        state["sim_first_price"] = close                  # 試撮後首價
        state["sim_high"]        = close
        state["sim_low"]         = close

        pre_price     = state["last_normal_price"]
        pre_type      = tick_type_str(state["last_normal_tick_type"])
        pre_vol       = state["last_normal_total_vol"]
        pre_price_str = f"{pre_price:.2f}" if pre_price is not None else "無前置"
        pct_str       = format_change_pct(change_pct)

        # 組合標籤
        tags = []
        if near_limit:
            tags.append(near_limit)
        if is_surge:
            tags.append("🚨大幅異動")
        tag_str = "　" + "　".join(tags) if tags else ""

        msg = (
            f"{code} 試撮:{close:.2f} 漲跌:{pct_str} 量:{tick.volume}張{tag_str}\n"
            f"前價:{pre_price_str} {pre_type} 累積量:{pre_vol}張"
        )
        print(f"🔥 【試撮警報】[{state['sim_start_time']}] {msg}")

        now = time.time()
        if code not in last_push_time or (now - last_push_time[code] > 60):
            # 大幅異動用不同標題讓手機更醒目
            if is_surge:
                bark_title = f"🚨大幅異動試撮 {code} {pct_str}"
            elif near_limit:
                bark_title = f"試撮警報　{near_limit}"
            else:
                bark_title = "試撮警報"
            send_bark_alert(bark_title, msg)
            last_push_time[code] = now

    else:
        state["sim_total_vol"] += tick.volume
        state["sim_price"]      = close
        if state["sim_high"] is None or close > state["sim_high"]:
            state["sim_high"] = close
        if state["sim_low"] is None or close < state["sim_low"]:
            state["sim_low"] = close
        if near_limit:
            state["near_limit"] = near_limit
        if change_pct is not None:
            state["change_pct"] = change_pct


# ─────────────── Excel 匯出 ───────────────────────────────

def export_to_excel() -> str:
    today_str  = datetime.now(TZ_TW).strftime("%Y%m%d")
    script_dir = os.path.dirname(os.path.abspath(__file__))
    filepath   = os.path.join(script_dir, f"試撮紀錄_{today_str}.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "試撮紀錄"

    headers    = ["日期", "股票代碼", "試撮開始", "試撮結束",
                  "進試撮前末價", "試撮後首價", "試撮末價", "結束後首價",
                  "漲跌幅%", "最後盤型", "試撮前累積量(張)", "試撮量(張)",
                  "最大利潤", "最大虧損", "漲跌停警示"]
    col_widths = [12, 10, 13, 13,
                  14, 12, 12, 12,
                  14, 10, 18, 12,
                  18, 18, 12]

    hdr_fill    = PatternFill("solid", fgColor="1F4E79")
    hdr_font    = Font(color="FFFFFF", bold=True, size=11)
    alt_fill    = PatternFill("solid", fgColor="DEEAF1")
    wht_fill    = PatternFill("solid", fgColor="FFFFFF")
    surge_fill  = PatternFill("solid", fgColor="FFE699")   # 大幅異動 → 黃底
    surge_font  = Font(color="C00000", bold=True)           # 深紅粗體

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell           = ws.cell(row=1, column=col_idx, value=h)
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.row_dimensions[1].height = 22

    for row_idx, r in enumerate(today_sim_records, 2):
        pct     = r.get("change_pct")
        pct_str = format_change_pct(pct)
        is_surge = pct is not None and abs(pct) >= SURGE_ALERT_PCT

        # 最大利潤/最大虧損：以「試撮後首價」為基準，比試撮期間最高/最低價
        base           = r.get("sim_first_price")
        max_profit_str = format_pl(r.get("sim_high"), base)
        max_loss_str   = format_pl(r.get("sim_low"),  base)

        values = [
            r.get("date", ""),       r["code"],                r["start_time"],          r["end_time"],
            r.get("pre_sim_price"),  r.get("sim_first_price"), r.get("sim_last_price"),  r["end_price"],
            pct_str,                 r["tick_type"],           r["pre_total_vol"],       r["sim_vol"],
            max_profit_str,          max_loss_str,             r["near_limit"],
        ]
        row_fill = alt_fill if row_idx % 2 == 0 else wht_fill

        for col_idx, val in enumerate(values, 1):
            cell           = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill      = row_fill
            cell.alignment = Alignment(horizontal="center")

        # 漲跌幅欄（第9欄）：大幅異動 → 黃底深紅
        pct_cell = ws.cell(row=row_idx, column=9)
        if is_surge:
            pct_cell.fill = surge_fill
            pct_cell.font = surge_font
        elif pct is not None:
            color = "C00000" if pct >= 0 else "375623"  # 上漲深紅 / 下跌深綠
            pct_cell.font = Font(color=color, bold=True)

        # 最大利潤（第13欄）紅、最大虧損（第14欄）綠（台股紅漲綠跌）
        ws.cell(row=row_idx, column=13).font = Font(color="C00000", bold=True)
        ws.cell(row=row_idx, column=14).font = Font(color="375623", bold=True)

        # 漲跌停警示欄（第15欄）標紅
        if r["near_limit"]:
            ws.cell(row=row_idx, column=15).font = Font(color="FF0000", bold=True)

    ws.freeze_panes = "A2"
    wb.save(filepath)
    print(f"\n📊 Excel 報表已儲存：{filepath}")
    return filepath


# ─────────────── 動態標的篩選 ────────────────────────────

def get_dynamic_market_list(api):
    official_excluded = []
    req_headers = {"User-Agent": "Mozilla/5.0"}

    try:
        urls = [
            "https://www.twse.com.tw/exchangeReport/TWTB4U?response=json",
            "https://www.twse.com.tw/exchangeReport/TWT11U?response=json",
        ]
        for url in urls:
            res  = requests.get(url, headers=req_headers, timeout=10)
            data = res.json()
            if "data" in data:
                official_excluded.extend([row[0].split(" ")[0] for row in data["data"]])
        print("📊 官方異常名單同步完成")
    except Exception as e:
        print(f"⚠️ 官方名單讀取失敗: {e}")

    target_categories = ["24","25","26","27","28","29","30","31","32","21","03","13","23"]
    MANUAL_BLACKLIST  = []

    candidate_contracts = []
    for contract in api.Contracts.Stocks.TSE:
        if contract.code in MANUAL_BLACKLIST or contract.code in official_excluded:
            continue
        if len(contract.code) != 4:
            continue
        if contract.category not in target_categories:
            continue
        if contract.day_trade != sj.constant.DayTrade.Yes:
            continue
        if hasattr(contract, "special_type") and contract.special_type != 0:
            continue
        candidate_contracts.append(contract)

    final_codes = []
    limit_info  = {}  # code -> (limit_up, limit_down, reference)
    # 漲跌停價、昨收參考價要從「合約 contract」取——snapshot 物件沒有這些欄位（先前 bug 來源）
    contract_by_code = {c.code: c for c in candidate_contracts}
    print(f"📈 正在分析 {len(candidate_contracts)} 檔標的的價格...")

    for i in range(0, len(candidate_contracts), 100):
        batch     = candidate_contracts[i:i+100]
        snapshots = api.snapshots(batch)
        for s in snapshots:
            if s.close and 15 <= s.close <= 300:
                final_codes.append(s.code)
                c = contract_by_code.get(s.code)
                # 昨收參考價：優先 contract.reference；取不到就用 snapshot「現價 − 漲跌額」反推
                ref = _to_float(getattr(c, "reference", None)) if c else None
                if ref is None:
                    chg = _to_float(getattr(s, "change_price", None))
                    if s.close is not None and chg is not None:
                        ref = round(float(s.close) - chg, 2)
                # 漲跌停價：優先 contract.limit_up/limit_down；取不到用昨收 ±10% 估
                lu = _to_float(getattr(c, "limit_up",   None)) if c else None
                ld = _to_float(getattr(c, "limit_down", None)) if c else None
                if lu is None and ref:
                    lu = round(ref * 1.1, 2)
                if ld is None and ref:
                    ld = round(ref * 0.9, 2)
                limit_info[s.code] = (lu, ld, ref)   # ← 新增 ref

    return final_codes[:254], limit_info


# ─────────────── 主程式 ───────────────────────────────────

def start_monitoring():
    api.login(api_key=API_KEY, secret_key=SECRET_KEY)
    api.on_tick_stk_v1()(on_tick_handler)   # 登入後再註冊 callback（新版 shioaji 需要）

    now_str = datetime.now(TZ_TW).strftime("%H:%M:%S")
    send_bark_alert("系統公告", f"監控程式已於 {now_str} 成功啟動！")
    print("✅ 登入成功！")

    final_monitor_list, limit_info = get_dynamic_market_list(api)
    print(f"🚀 啟動監控！實際監控標的：{len(final_monitor_list)} 檔")

    for code in final_monitor_list:
        lu, ld, ref = limit_info.get(code, (None, None, None))
        _init_state(code, limit_up=lu, limit_down=ld, reference=ref)

    for code in final_monitor_list:
        contract = api.Contracts.Stocks[code]
        api.quote.subscribe(contract, quote_type=sj.constant.QuoteType.Tick)

    last_heartbeat_time = 0
    try:
        while True:
            now    = time.time()
            tw_now = datetime.now(TZ_TW)

            if (tw_now.hour > MARKET_CLOSE_HOUR or
                    (tw_now.hour == MARKET_CLOSE_HOUR and tw_now.minute >= MARKET_CLOSE_MINUTE)):
                print(f"\n🔔 [{tw_now.strftime('%H:%M:%S')}] 已過收盤時間，自動結束監控")
                raise SystemExit(0)

            if now - last_heartbeat_time >= 300:
                print(
                    f"💓 [心跳] {tw_now.strftime('%H:%M:%S')} "
                    f"監控中... 已記錄 {len(today_sim_records)} 筆試撮"
                )
                last_heartbeat_time = now
            time.sleep(1)

    except (KeyboardInterrupt, SystemExit):
        print("\n👋 監控結束，正在匯出報表...")
        if today_sim_records:
            filepath = export_to_excel()
            send_bark_alert("試撮報表", f"今日記錄 {len(today_sim_records)} 筆，報表已儲存")
        else:
            print("📭 今日無試撮紀錄")
            send_bark_alert("試撮監控", "今日無符合條件的試撮紀錄")
        print("正在登出...")
        api.logout()


if __name__ == "__main__":
    start_monitoring()
